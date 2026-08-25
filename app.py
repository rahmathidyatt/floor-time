"""
Brighton Floor Time Scheduler
=============================

Aplikasi Streamlit untuk menyiapkan data agen bulan sebelumnya, mengatur hari
libur, membuat jadwal Floor Time satu bulan secara adil, memvalidasi hasil,
serta mengekspor poster mingguan ke Excel, PNG, dan PDF.

Jalankan:
    pip install -r requirements.txt
    streamlit run app.py
"""

from __future__ import annotations

import calendar
import hashlib
import html
import io
import os
import random
import re
import zipfile
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from math import ceil
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont


# -----------------------------------------------------------------------------
# KONFIGURASI DASAR
# -----------------------------------------------------------------------------

APP_TITLE = "Brighton Floor Time Scheduler"
DEFAULT_HUB_NAME = "HUB CIBUBUR"
DEFAULT_LOGO_TEXT = "Brighton"
EXPORT_RENDER_SCALE = 2

BRIGHTON_YELLOW = "#FFD10A"
BRIGHTON_BLACK = "#111111"
HOLIDAY_RED = "#D92D20"
HOLIDAY_BG = "#FFF7F6"
TEXT_SECONDARY = "#515151"
BORDER_COLOR = "#D9D9D9"
SOFT_GRAY = "#F5F5F5"

MONTH_NAMES_ID = {
    1: "Januari", 2: "Februari", 3: "Maret", 4: "April",
    5: "Mei", 6: "Juni", 7: "Juli", 8: "Agustus",
    9: "September", 10: "Oktober", 11: "November", 12: "Desember",
}

DAY_NAMES_ID = {
    0: "SENIN", 1: "SELASA", 2: "RABU", 3: "KAMIS",
    4: "JUMAT", 5: "SABTU", 6: "MINGGU",
}

SHIFT_TEMPLATES = {
    0: [("Pagi", "08.00-12.30"), ("Siang", "12.30-17.00")],
    1: [("Pagi", "08.00-12.30"), ("Siang", "12.30-17.00")],
    2: [("Pagi", "08.00-12.30"), ("Siang", "12.30-17.00")],
    3: [("Pagi", "08.00-12.30"), ("Siang", "12.30-17.00")],
    4: [("Pagi", "08.00-12.30"), ("Siang", "12.30-17.00")],
    5: [("Pagi", "08.00-11.30"), ("Siang", "11.30-15.00")],
}
SHIFT_NAME_TO_INDEX = {"Pagi": 0, "Siang": 1}


# -----------------------------------------------------------------------------
# DATA MODEL
# -----------------------------------------------------------------------------

@dataclass(frozen=True)
class AgentRecord:
    agent_id: str
    name: str
    code: str
    business_unit: str
    previous_month_attendance: int
    job_title: str = ""

    @property
    def display_name(self) -> str:
        return f"{self.name} ({self.code})" if self.code else self.name

    @property
    def office(self) -> str:
        """Office disimpan pada field business_unit untuk backward compatibility."""
        return self.business_unit

    @property
    def unit(self) -> str:
        """Ambil nama unit dari kata kedua pada kolom Office.

        Contoh: ``Brighton Priority Cibubur, Bogor`` -> ``Priority``.
        Jika data Office hanya memiliki satu kata, nilai tersebut dipakai sebagai
        fallback agar informasi tidak hilang.
        """
        words = clean_text(self.office).split()
        if not words:
            return ""
        unit = words[1] if len(words) >= 2 else words[0]
        return unit.strip(" ,.;:-")


@dataclass(frozen=True)
class HolidayInfo:
    tanggal: date
    name: str


@dataclass(frozen=True)
class ShiftSlot:
    week_no: int
    tanggal: date
    day_name: str
    shift_index: int
    shift_name: str
    time_label: str
    base_capacity: int

    @property
    def key(self) -> str:
        return f"{self.tanggal.isoformat()}__{self.shift_index}"


@dataclass(frozen=True)
class UrgentRequest:
    agent_id: str
    tanggal: date
    shift_index: int
    note: str = ""


@dataclass(frozen=True)
class AssignmentEntry:
    agent_id: str
    display_name: str
    is_urgent: bool = False


# -----------------------------------------------------------------------------
# FORMAT & UTILITAS
# -----------------------------------------------------------------------------

def format_date_id(d: date, include_year: bool = False) -> str:
    if include_year:
        return f"{d.day:02d} {MONTH_NAMES_ID[d.month]} {d.year}"
    return f"{d.day:02d} {MONTH_NAMES_ID[d.month]}"


def month_label(year: int, month: int) -> str:
    return f"{MONTH_NAMES_ID[month]} {year}"


def previous_month(year: int, month: int) -> Tuple[int, int]:
    if month == 1:
        return year - 1, 12
    return year, month - 1


def date_option_label(d: date) -> str:
    return f"{DAY_NAMES_ID[d.weekday()]} | {format_date_id(d, include_year=True)}"


def week_range_label(week_dates: Sequence[date]) -> str:
    if not week_dates:
        return ""
    start = min(week_dates)
    end = max(week_dates)
    if start.month == end.month and start.year == end.year:
        return f"{start.day:02d} {MONTH_NAMES_ID[start.month].upper()} {start.year} - {end.day:02d} {MONTH_NAMES_ID[end.month].upper()} {end.year}"
    return f"{format_date_id(start, True).upper()} - {format_date_id(end, True).upper()}"


def clean_text(value: object) -> str:
    text = "" if value is None else str(value)
    if text.lower() == "nan":
        return ""
    text = text.strip().strip("-•*;")
    return re.sub(r"\s+", " ", text).strip()


def html_escape(value: object) -> str:
    return html.escape("" if value is None else str(value))


def parse_agent_display(value: object) -> Tuple[str, str]:
    text = clean_text(value)
    if not text:
        return "", ""
    match = re.match(r"^(.*?)\s*\(([^()]*)\)\s*$", text)
    if match:
        return clean_text(match.group(1)), clean_text(match.group(2)).upper()
    return text, ""


def make_agent_id(name: str, code: str) -> str:
    raw = f"{name.casefold()}|{code.casefold()}".encode("utf-8")
    return hashlib.sha1(raw).hexdigest()[:14]


def parse_int_nonnegative(value: object) -> int:
    if value in (None, ""):
        return 0
    try:
        return max(0, int(float(value)))
    except (TypeError, ValueError):
        return 0


def dataframe_to_agents(df: pd.DataFrame) -> Tuple[List[AgentRecord], List[str]]:
    """Normalisasi tabel agen dan pertahankan kompatibilitas format lama."""
    if df is None or df.empty:
        return [], []

    normalized_cols = {str(c).strip().casefold(): c for c in df.columns}

    def find_col(candidates: Sequence[str]) -> Optional[object]:
        for candidate in candidates:
            key = candidate.casefold()
            if key in normalized_cols:
                return normalized_cols[key]
        return None

    agent_col = find_col(["Agent", "Agen", "Nama Agen", "Nama", "Property Advisor", "PA"])
    name_col = find_col(["Name"])
    code_col = find_col(["Code", "Kode", "Kode Agen"])
    attendance_col = find_col([
        "Kehadiran", "Attendance", "Previous Month Attendance",
        "previousMonthAttendance", "Total Kehadiran", "Total",
    ])
    bu_col = find_col([
        "Office", "Kantor", "Business Unit", "BusinessUnit",
        "Bisnis Unit", "BU", "Unit",
    ])
    job_title_col = find_col(["Jabatan", "Position", "Role", "Job Title", "Title"])

    if agent_col is None and name_col is None:
        agent_col = df.columns[0]

    agents: List[AgentRecord] = []
    warnings: List[str] = []
    seen_ids = set()

    for row_idx, row in df.iterrows():
        if name_col is not None:
            name = clean_text(row.get(name_col, ""))
            code = clean_text(row.get(code_col, "")).upper() if code_col is not None else ""
        else:
            name, parsed_code = parse_agent_display(row.get(agent_col, ""))
            explicit_code = clean_text(row.get(code_col, "")).upper() if code_col is not None else ""
            code = explicit_code or parsed_code

        if not name:
            continue

        agent_id = make_agent_id(name, code)
        if agent_id in seen_ids:
            warnings.append(f"Duplikat agen diabaikan: {name}{f' ({code})' if code else ''}.")
            continue
        seen_ids.add(agent_id)

        business_unit = clean_text(row.get(bu_col, "")) if bu_col is not None else ""
        job_title = clean_text(row.get(job_title_col, "")) if job_title_col is not None else ""
        attendance = parse_int_nonnegative(row.get(attendance_col, 0)) if attendance_col is not None else 0
        agents.append(
            AgentRecord(
                agent_id=agent_id,
                name=name,
                code=code,
                business_unit=business_unit,
                previous_month_attendance=attendance,
                job_title=job_title,
            )
        )

    return agents, warnings


def filter_agents_for_schedule(
    agents: Sequence[AgentRecord],
    minimum_attendance: int = 0,
    excluded_agent_ids: Optional[Iterable[str]] = None,
) -> List[AgentRecord]:
    """Pilih agen aktif berdasarkan minimum kehadiran dan pengecualian manual.

    Filter ini hanya memengaruhi jadwal bulan yang sedang disusun. Data asli hasil
    upload tidak diubah, sehingga user dapat mengubah ambang atau mengaktifkan
    kembali agen tanpa mengunggah ulang file.
    """
    minimum = max(0, int(minimum_attendance))
    excluded = set(excluded_agent_ids or [])
    return [
        agent
        for agent in agents
        if agent.previous_month_attendance >= minimum and agent.agent_id not in excluded
    ]


def sort_agents_for_roster(
    agents: Sequence[AgentRecord],
    sort_mode: str = "Default (sesuai file)",
) -> List[AgentRecord]:
    """Urutkan daftar agen untuk poster bulanan tanpa mengubah data sumber."""
    ordered = list(agents)
    if sort_mode == "Terbanyak ke terendah":
        return sorted(ordered, key=lambda a: -a.previous_month_attendance)
    if sort_mode == "Terendah ke terbanyak":
        return sorted(ordered, key=lambda a: a.previous_month_attendance)
    return ordered


def read_agent_upload(uploaded_file) -> pd.DataFrame:
    if uploaded_file is None:
        return pd.DataFrame(columns=["Nama", "Jabatan", "Office", "Total Kehadiran"])
    filename = uploaded_file.name.lower()
    if filename.endswith(".csv"):
        return pd.read_csv(uploaded_file)
    if filename.endswith((".xlsx", ".xls")):
        return pd.read_excel(uploaded_file)
    raise ValueError("Format file belum didukung. Gunakan CSV atau Excel.")


# -----------------------------------------------------------------------------
# KALENDER & SLOT
# -----------------------------------------------------------------------------

def get_month_work_dates(year: int, month: int) -> List[date]:
    """Ambil Senin-Sabtu dari minggu yang menyentuh bulan target.

    Dengan aturan ini September 2026 dimulai dari Senin 31 Agustus 2026 dan
    minggu terakhir diteruskan sampai Sabtu, sehingga cross-month konsisten.
    """
    first_day = date(year, month, 1)
    last_day = date(year, month, calendar.monthrange(year, month)[1])
    start_monday = first_day - timedelta(days=first_day.weekday())

    last_active = last_day
    if last_active.weekday() == 6:
        last_active -= timedelta(days=1)
    end_saturday = last_active + timedelta(days=(5 - last_active.weekday()))

    dates: List[date] = []
    current = start_monday
    while current <= end_saturday:
        if current.weekday() <= 5:
            dates.append(current)
        current += timedelta(days=1)
    return dates


def group_dates_by_calendar_week(work_dates: Iterable[date]) -> Dict[int, List[date]]:
    grouped: Dict[date, List[date]] = {}
    for d in work_dates:
        monday = d - timedelta(days=d.weekday())
        grouped.setdefault(monday, []).append(d)
    return {idx: sorted(grouped[monday]) for idx, monday in enumerate(sorted(grouped), start=1)}


def build_slots_by_week(
    weeks: Dict[int, List[date]],
    holidays: Dict[date, HolidayInfo],
    weekday_capacity: int,
    saturday_capacity: int,
) -> Dict[int, List[ShiftSlot]]:
    slots_by_week: Dict[int, List[ShiftSlot]] = {}
    for week_no, dates_in_week in weeks.items():
        slots: List[ShiftSlot] = []
        for d in dates_in_week:
            if d.weekday() > 5 or d in holidays:
                continue
            capacity = saturday_capacity if d.weekday() == 5 else weekday_capacity
            for shift_index, (shift_name, time_label) in enumerate(SHIFT_TEMPLATES[d.weekday()]):
                slots.append(
                    ShiftSlot(
                        week_no=week_no,
                        tanggal=d,
                        day_name=DAY_NAMES_ID[d.weekday()],
                        shift_index=shift_index,
                        shift_name=shift_name,
                        time_label=time_label,
                        base_capacity=capacity,
                    )
                )
        slots_by_week[week_no] = slots
    return slots_by_week


def build_holidays(selected_labels: Sequence[str], label_to_date: Dict[str, date]) -> Dict[date, HolidayInfo]:
    holidays: Dict[date, HolidayInfo] = {}
    for label in selected_labels:
        d = label_to_date[label]
        key = f"holiday_name_{d.isoformat()}"
        holiday_name = clean_text(st.session_state.get(key, "")) or "Hari Libur"
        holidays[d] = HolidayInfo(d, holiday_name)
    return holidays


# -----------------------------------------------------------------------------
# REQUEST KHUSUS
# -----------------------------------------------------------------------------

def parse_urgent_requests(
    urgent_df: pd.DataFrame,
    label_to_date: Dict[str, date],
    agents: Sequence[AgentRecord],
    holidays: Dict[date, HolidayInfo],
) -> Tuple[List[UrgentRequest], List[str]]:
    requests: List[UrgentRequest] = []
    errors: List[str] = []
    display_to_agent = {a.display_name: a for a in agents}

    if urgent_df is None or urgent_df.empty:
        return requests, errors

    for idx, row in urgent_df.iterrows():
        display = clean_text(row.get("Agen", ""))
        date_label = clean_text(row.get("Tanggal", ""))
        shift_name = clean_text(row.get("Shift", ""))
        note = clean_text(row.get("Catatan", ""))
        if not any([display, date_label, shift_name, note]):
            continue
        row_no = idx + 1
        if display not in display_to_agent:
            errors.append(f"Request baris {row_no}: agen tidak valid.")
            continue
        if date_label not in label_to_date:
            errors.append(f"Request baris {row_no}: tanggal belum valid.")
            continue
        d = label_to_date[date_label]
        if d in holidays:
            errors.append(f"Request baris {row_no}: {format_date_id(d, True)} adalah hari libur.")
            continue
        if shift_name not in SHIFT_NAME_TO_INDEX:
            errors.append(f"Request baris {row_no}: shift harus Pagi atau Siang.")
            continue
        requests.append(UrgentRequest(display_to_agent[display].agent_id, d, SHIFT_NAME_TO_INDEX[shift_name], note))
    return requests, errors


# -----------------------------------------------------------------------------
# MESIN SCHEDULING
# -----------------------------------------------------------------------------

def make_rng(seed_text: Optional[str]) -> random.Random:
    if seed_text and seed_text.strip():
        return random.Random(seed_text.strip())
    return random.SystemRandom()


def recommended_weekly_frequency(agent_count: int, total_capacity: int, active_days: int) -> Optional[int]:
    if agent_count <= 0:
        return None
    minimum = ceil(total_capacity / agent_count)
    minimum = min(minimum, active_days)
    return minimum


def capacity_analysis(
    agents: Sequence[AgentRecord],
    slots_by_week: Dict[int, List[ShiftSlot]],
    max_weekly_assignments: int,
) -> List[str]:
    messages: List[str] = []
    if not agents:
        return messages
    for week_no, slots in slots_by_week.items():
        if not slots:
            continue
        total_capacity = sum(s.base_capacity for s in slots)
        active_days = len({s.tanggal for s in slots})
        max_possible = len(agents) * min(max_weekly_assignments, active_days)
        if max_possible < total_capacity:
            recommended = recommended_weekly_frequency(len(agents), total_capacity, active_days)
            if recommended is not None and recommended <= 3:
                messages.append(
                    f"Periode {week_range_label(sorted({s.tanggal for s in slots}))}: kapasitas {total_capacity} assignment, "
                    f"tetapi konfigurasi saat ini maksimal {max_possible}. Rekomendasi minimum sekitar {recommended}x per agen/minggu."
                )
            else:
                messages.append(
                    f"Periode {week_range_label(sorted({s.tanggal for s in slots}))}: kapasitas {total_capacity} assignment melebihi "
                    f"kapasitas maksimum {max_possible}. Sebagian slot dapat tetap kosong."
                )
    return messages


def _slot_score(
    slot: ShiftSlot,
    schedule_for_week: Dict[str, List[AssignmentEntry]],
    capacity_for_week: Dict[str, int],
    agent: AgentRecord,
    agent_dates: Dict[str, set],
    month_shift_counts: Dict[str, Dict[int, int]],
    last_date_by_agent: Dict[str, Optional[date]],
    agent_map: Dict[str, AgentRecord],
    rng: random.Random,
) -> Tuple[float, int, int, int, float]:
    load_ratio = len(schedule_for_week[slot.key]) / max(capacity_for_week[slot.key], 1)
    shift_repeat = month_shift_counts[agent.agent_id][slot.shift_index]
    same_business_unit = sum(
        1
        for entry in schedule_for_week[slot.key]
        if agent.business_unit
        and agent_map.get(entry.agent_id) is not None
        and agent_map[entry.agent_id].business_unit == agent.business_unit
    )
    last_date = last_date_by_agent.get(agent.agent_id)
    adjacent_penalty = 0
    if last_date is not None:
        delta = abs((slot.tanggal - last_date).days)
        adjacent_penalty = 2 if delta == 1 else (1 if delta == 2 else 0)
    return load_ratio, same_business_unit, adjacent_penalty, shift_repeat, rng.random()


def generate_schedule(
    agents: Sequence[AgentRecord],
    slots_by_week: Dict[int, List[ShiftSlot]],
    urgent_requests: Sequence[UrgentRequest],
    max_weekly_assignments: int,
    auto_expand_capacity: bool,
    seed_text: Optional[str],
) -> Tuple[pd.DataFrame, Dict[int, Dict[str, List[AssignmentEntry]]], Dict[int, Dict[str, int]], List[str]]:
    """Generate schedule fair-distribution dengan aturan unik per tanggal.

    - Semua agen diprioritaskan memperoleh minimal 1 assignment/minggu jika slot cukup.
    - Jika target kapasitas masih belum penuh dan maxWeekly > 1, agen dapat berulang
      pada hari berbeda secara merata hingga batas 2x/3x.
    - Agen tidak pernah boleh muncul dua kali pada tanggal yang sama.
    """
    rng = make_rng(seed_text)
    warnings: List[str] = []
    schedule: Dict[int, Dict[str, List[AssignmentEntry]]] = {}
    capacity_by_week: Dict[int, Dict[str, int]] = {}
    rows: List[dict] = []

    agent_map = {a.agent_id: a for a in agents}
    month_counts = {a.agent_id: 0 for a in agents}
    month_shift_counts = {a.agent_id: {0: 0, 1: 0} for a in agents}
    last_date_by_agent: Dict[str, Optional[date]] = {a.agent_id: None for a in agents}

    date_to_week: Dict[date, int] = {}
    slot_lookup: Dict[int, Dict[Tuple[date, int], ShiftSlot]] = {}
    for week_no, slots in slots_by_week.items():
        slot_lookup[week_no] = {}
        for slot in slots:
            date_to_week[slot.tanggal] = week_no
            slot_lookup[week_no][(slot.tanggal, slot.shift_index)] = slot

    urgent_by_week: Dict[int, List[UrgentRequest]] = {}
    for req in urgent_requests:
        week_no = date_to_week.get(req.tanggal)
        if week_no is None:
            warnings.append(f"Request {agent_map.get(req.agent_id, AgentRecord('', 'Agen', '', '', 0)).display_name} dilewati karena tanggal tidak aktif.")
            continue
        urgent_by_week.setdefault(week_no, []).append(req)

    for week_no, slots in slots_by_week.items():
        schedule[week_no] = {slot.key: [] for slot in slots}
        capacity_by_week[week_no] = {slot.key: slot.base_capacity for slot in slots}
        if not slots:
            warnings.append(f"Periode {week_no} tidak memiliki hari kerja aktif.")
            continue

        total_initial_capacity = sum(s.base_capacity for s in slots)
        if len(agents) > total_initial_capacity and auto_expand_capacity:
            target_per_slot = ceil(len(agents) / len(slots))
            for slot in slots:
                capacity_by_week[week_no][slot.key] = max(slot.base_capacity, target_per_slot)
            warnings.append(
                f"Periode {week_range_label(sorted({s.tanggal for s in slots}))}: kapasitas shift diperluas agar seluruh agen tetap bisa mendapat minimal satu jadwal."
            )

        weekly_counts = {a.agent_id: 0 for a in agents}
        agent_dates: Dict[str, set] = {a.agent_id: set() for a in agents}

        def assign(agent: AgentRecord, slot: ShiftSlot, urgent: bool = False) -> bool:
            if slot.tanggal in agent_dates[agent.agent_id]:
                return False
            if weekly_counts[agent.agent_id] >= max_weekly_assignments:
                return False
            if len(schedule[week_no][slot.key]) >= capacity_by_week[week_no][slot.key]:
                if urgent and auto_expand_capacity:
                    capacity_by_week[week_no][slot.key] += 1
                else:
                    return False
            schedule[week_no][slot.key].append(AssignmentEntry(agent.agent_id, agent.display_name, urgent))
            weekly_counts[agent.agent_id] += 1
            agent_dates[agent.agent_id].add(slot.tanggal)
            month_counts[agent.agent_id] += 1
            month_shift_counts[agent.agent_id][slot.shift_index] += 1
            last_date_by_agent[agent.agent_id] = slot.tanggal
            return True

        # 1) Request khusus ditempatkan lebih dahulu, tetap tunduk pada integrity rules.
        for req in urgent_by_week.get(week_no, []):
            agent = agent_map.get(req.agent_id)
            slot = slot_lookup[week_no].get((req.tanggal, req.shift_index))
            if agent is None or slot is None:
                warnings.append("Satu request khusus dilewati karena agen/slot tidak valid.")
                continue
            if not assign(agent, slot, urgent=True):
                warnings.append(
                    f"Request {agent.display_name} pada {format_date_id(req.tanggal, True)} dilewati karena bentrok tanggal, batas mingguan, atau kapasitas."
                )

        # 2) Prioritaskan setiap agen memperoleh minimal satu jadwal per minggu.
        ordered_agents = list(agents)
        rng.shuffle(ordered_agents)
        ordered_agents.sort(key=lambda a: (month_counts[a.agent_id], -a.previous_month_attendance))

        for agent in ordered_agents:
            if weekly_counts[agent.agent_id] > 0:
                continue
            candidates = [
                s for s in slots
                if s.tanggal not in agent_dates[agent.agent_id]
                and len(schedule[week_no][s.key]) < capacity_by_week[week_no][s.key]
            ]
            if not candidates and auto_expand_capacity:
                # Untuk kasus agen lebih banyak daripada slot capacity, perluas slot paling ringan.
                candidates = sorted(slots, key=lambda s: len(schedule[week_no][s.key]))[:1]
                if candidates:
                    capacity_by_week[week_no][candidates[0].key] += 1
            if not candidates:
                warnings.append(f"{agent.display_name} belum mendapat jadwal pada {week_range_label(sorted({s.tanggal for s in slots}))} karena kapasitas tidak cukup.")
                continue
            candidates.sort(
                key=lambda s: _slot_score(
                    s, schedule[week_no], capacity_by_week[week_no], agent,
                    agent_dates, month_shift_counts, last_date_by_agent, agent_map, rng,
                )
            )
            assign(agent, candidates[0])

        # 3) Isi sisa kapasitas secara merata jika repeat 2x/3x diizinkan.
        while True:
            open_slots = [s for s in slots if len(schedule[week_no][s.key]) < capacity_by_week[week_no][s.key]]
            if not open_slots:
                break

            # Mulai dari slot dengan rasio isi paling rendah agar distribusi antar shift seimbang.
            open_slots.sort(
                key=lambda s: (
                    len(schedule[week_no][s.key]) / max(capacity_by_week[week_no][s.key], 1),
                    s.tanggal,
                    s.shift_index,
                )
            )
            progress = False
            for slot in open_slots:
                eligible = [
                    a for a in agents
                    if weekly_counts[a.agent_id] < max_weekly_assignments
                    and slot.tanggal not in agent_dates[a.agent_id]
                ]
                if not eligible:
                    continue
                rng.shuffle(eligible)
                eligible.sort(
                    key=lambda a: (
                        weekly_counts[a.agent_id],
                        month_counts[a.agent_id],
                        sum(
                            1 for entry in schedule[week_no][slot.key]
                            if a.business_unit
                            and agent_map.get(entry.agent_id) is not None
                            and agent_map[entry.agent_id].business_unit == a.business_unit
                        ),
                        month_shift_counts[a.agent_id][slot.shift_index],
                        1 if last_date_by_agent[a.agent_id] and abs((slot.tanggal - last_date_by_agent[a.agent_id]).days) == 1 else 0,
                        -a.previous_month_attendance,
                    )
                )
                if assign(eligible[0], slot):
                    progress = True
            if not progress:
                break

        unfilled = sum(
            max(0, capacity_by_week[week_no][slot.key] - len(schedule[week_no][slot.key]))
            for slot in slots
        )
        if unfilled:
            warnings.append(
                f"Periode {week_range_label(sorted({s.tanggal for s in slots}))}: {unfilled} slot kapasitas belum terisi karena batas frekuensi/tanggal unik tercapai."
            )

        for slot in slots:
            for order_no, entry in enumerate(schedule[week_no][slot.key], start=1):
                agent = agent_map[entry.agent_id]
                rows.append(
                    {
                        "Minggu Internal": week_no,
                        "Tanggal": slot.tanggal.isoformat(),
                        "Hari": slot.day_name,
                        "Shift": slot.shift_name,
                        "Jam": slot.time_label,
                        "Urutan": order_no,
                        "Agent ID": agent.agent_id,
                        "Agen": agent.display_name,
                        "Kode": agent.code,
                        "Jabatan": agent.job_title,
                        "Unit": agent.unit,
                        "Kehadiran Bulan Sebelumnya": agent.previous_month_attendance,
                        "Urgent": "Ya" if entry.is_urgent else "Tidak",
                    }
                )

    return pd.DataFrame(rows), schedule, capacity_by_week, warnings


def validate_schedule(
    assignments_df: pd.DataFrame,
    agents: Sequence[AgentRecord],
    weeks: Dict[int, List[date]],
    holidays: Dict[date, HolidayInfo],
    max_weekly_assignments: int,
) -> Tuple[bool, List[str]]:
    messages: List[str] = []
    if assignments_df.empty:
        return False, ["Jadwal kosong. Periksa daftar agen, hari aktif, dan kapasitas shift."]

    valid = True
    agent_ids = {a.agent_id for a in agents}
    allowed_dates = {d for dates in weeks.values() for d in dates}
    holiday_dates = set(holidays)

    df = assignments_df.copy()
    df["_date"] = pd.to_datetime(df["Tanggal"], errors="coerce").dt.date

    invalid_agents = df[~df["Agent ID"].isin(agent_ids)]
    if not invalid_agents.empty:
        valid = False
        messages.append("Terdapat assignment dengan Agent ID yang tidak dikenal.")

    if df["_date"].isna().any() or any(d not in allowed_dates for d in df["_date"].dropna()):
        valid = False
        messages.append("Terdapat tanggal assignment di luar periode jadwal.")

    if any(d in holiday_dates for d in df["_date"].dropna()):
        valid = False
        messages.append("Terdapat agen yang masih terjadwal pada hari libur.")

    duplicate_date = df.groupby(["Agent ID", "_date"]).size()
    duplicate_date = duplicate_date[duplicate_date > 1]
    if not duplicate_date.empty:
        valid = False
        messages.append("Terdapat agen yang muncul lebih dari sekali pada tanggal yang sama.")

    duplicate_record = df.duplicated(subset=["Agent ID", "Tanggal", "Shift"], keep=False)
    if duplicate_record.any():
        valid = False
        messages.append("Terdapat duplicate assignment record pada agen/tanggal/shift yang sama.")

    if not set(df["Shift"].dropna().unique()).issubset(set(SHIFT_NAME_TO_INDEX)):
        valid = False
        messages.append("Terdapat nama shift yang tidak valid.")

    for week_no, dates_in_week in weeks.items():
        week_df = df[df["Minggu Internal"] == week_no]
        counts = week_df.groupby("Agent ID").size() if not week_df.empty else pd.Series(dtype=int)
        over = counts[counts > max_weekly_assignments]
        if not over.empty:
            valid = False
            names = [next((a.display_name for a in agents if a.agent_id == aid), aid) for aid in over.index]
            messages.append(
                f"Periode {week_range_label(dates_in_week)}: batas {max_weekly_assignments}x/minggu terlampaui oleh {', '.join(names)}."
            )

    return valid, messages


# -----------------------------------------------------------------------------
# HTML POSTER (fallback print-friendly)
# -----------------------------------------------------------------------------

def build_schedule_css() -> str:
    return f"""
    <style>
        :root {{
            --brighton-yellow: {BRIGHTON_YELLOW};
            --brighton-black: {BRIGHTON_BLACK};
            --holiday-red: {HOLIDAY_RED};
            --holiday-bg: {HOLIDAY_BG};
            --border: {BORDER_COLOR};
            --muted: {TEXT_SECONDARY};
        }}
        * {{ box-sizing: border-box; }}
        body {{ margin:0; font-family: Inter, Arial, Helvetica, sans-serif; color:var(--brighton-black); background:#fff; }}
        .poster {{ width:100%; max-width:1080px; margin:0 auto 28px; padding:36px 48px 24px; background:#fff; page-break-after:always; }}
        .brand {{ text-align:center; font-size:62px; font-weight:800; letter-spacing:-2px; }}
        .brand .o {{ color:var(--brighton-yellow); }}
        .hub {{ text-align:center; font-size:18px; font-weight:800; letter-spacing:4px; margin-top:3px; }}
        .brand-lines {{ display:flex; justify-content:center; gap:20px; margin:12px 0 26px; }}
        .brand-lines span {{ width:60px; height:3px; background:var(--brighton-yellow); }}
        .title {{ text-align:center; font-size:42px; font-weight:900; margin:0; letter-spacing:.5px; }}
        .title-line {{ display:none; }}
        .range {{ text-align:center; font-size:22px; font-weight:800; margin-bottom:26px; }}
        .day-card {{ display:grid; grid-template-columns:25% 37.5% 37.5%; border:1px solid var(--border); border-radius:10px; margin:0 0 10px; overflow:hidden; min-height:138px; }}
        .day-meta {{ padding:24px 24px; border-right:1px solid #eee; }}
        .day-name {{ font-size:27px; font-weight:900; }}
        .day-date {{ font-size:17px; margin-top:9px; }}
        .day-accent {{ width:42px; height:2px; background:var(--brighton-yellow); margin-top:16px; }}
        .shift {{ padding:22px 24px; }}
        .shift + .shift {{ border-left:1px dashed #ddd; }}
        .time {{ display:flex; gap:10px; align-items:center; font-size:17px; font-weight:800; margin-bottom:12px; }}
        .clock {{ width:24px; height:24px; border:2px solid var(--brighton-yellow); border-radius:50%; display:inline-block; }}
        .agents {{ margin:0; padding-left:20px; font-size:15px; line-height:1.55; }}
        .special {{ float:right; background:var(--brighton-yellow); padding:4px 9px; border-radius:5px; font-size:11px; font-weight:800; }}
        .holiday {{ border-color:#ef9a94; background:var(--holiday-bg); grid-template-columns:25% 75%; }}
        .holiday .day-meta {{ border-right:1px solid #ef9a94; }}
        .holiday .day-name, .holiday .day-date, .holiday .holiday-name {{ color:var(--holiday-red); }}
        .holiday .day-accent {{ background:var(--holiday-red); }}
        .holiday-name {{ display:flex; align-items:center; padding:24px 38px; font-size:19px; font-weight:700; }}
        .notes {{ margin-top:14px; border:1px solid var(--border); border-radius:10px; padding:20px 26px; display:grid; grid-template-columns:150px 1fr; gap:20px; }}
        .notes-title {{ font-size:19px; font-weight:900; border-right:3px solid var(--brighton-yellow); }}
        .notes ul {{ margin:0; padding-left:22px; font-size:14px; line-height:1.65; }}
        @media print {{ .poster {{ max-width:none; margin:0; border:none; }} }}
    </style>
    """


def render_agent_list_html(entries: Sequence[AssignmentEntry]) -> str:
    if not entries:
        return "<ul class='agents'><li>Slot Belum Terisi</li></ul>"
    return "<ul class='agents'>" + "".join(f"<li>{html_escape(e.display_name)}</li>" for e in entries) + "</ul>"


def render_week_html(
    week_dates: Sequence[date],
    slots: Sequence[ShiftSlot],
    schedule_for_week: Dict[str, List[AssignmentEntry]],
    holidays: Dict[date, HolidayInfo],
    hub_name: str,
    logo_text: str,
    notes: Sequence[str],
) -> str:
    slot_by_date: Dict[date, List[ShiftSlot]] = {}
    for slot in slots:
        slot_by_date.setdefault(slot.tanggal, []).append(slot)

    cards: List[str] = []
    for d in sorted(week_dates):
        if d.weekday() > 5:
            continue
        if d in holidays:
            cards.append(f"""
            <section class="day-card holiday">
              <div class="day-meta"><div class="day-name">{DAY_NAMES_ID[d.weekday()]}</div><div class="day-date">{format_date_id(d)}</div><div class="day-accent"></div></div>
              <div class="holiday-name">{html_escape(holidays[d].name)}</div>
            </section>""")
            continue

        day_slots = sorted(slot_by_date.get(d, []), key=lambda s: s.shift_index)
        shift_html = ""
        for slot in day_slots:
            badge = "<span class='special'>Jam Operasional Khusus</span>" if d.weekday() == 5 and slot.shift_index == 1 else ""
            shift_html += f"""
              <div class="shift">{badge}<div class="time"><span class="clock"></span>{html_escape(slot.time_label)}</div>{render_agent_list_html(schedule_for_week.get(slot.key, []))}</div>
            """
        cards.append(f"""
        <section class="day-card">
          <div class="day-meta"><div class="day-name">{DAY_NAMES_ID[d.weekday()]}</div><div class="day-date">{format_date_id(d)}</div><div class="day-accent"></div></div>
          {shift_html}
        </section>""")

    safe_logo = html_escape(logo_text or DEFAULT_LOGO_TEXT)
    if safe_logo.casefold() == "brighton":
        safe_logo = "Bright<span class='o'>o</span>n"
    notes_html = "".join(f"<li>{html_escape(n)}</li>" for n in notes if clean_text(n))
    return f"""
    <div class="poster">
      <div class="brand">{safe_logo}</div>
      <div class="hub">{html_escape(hub_name)}</div>
      <div class="brand-lines"><span></span><span></span></div>
      <h1 class="title">FLOOR TIME SCHEDULE</h1>
      <div class="range" style="margin-top:18px;">{week_range_label(week_dates)}</div>
      {''.join(cards)}
      <div class="notes"><div class="notes-title">CATATAN</div><ul>{notes_html}</ul></div>
    </div>
    """


def build_full_html_document(
    weeks: Dict[int, List[date]],
    slots_by_week: Dict[int, List[ShiftSlot]],
    schedule: Dict[int, Dict[str, List[AssignmentEntry]]],
    holidays: Dict[date, HolidayInfo],
    hub_name: str,
    logo_text: str,
    notes: Sequence[str],
) -> str:
    pages = [
        render_week_html(
            week_dates=week_dates,
            slots=slots_by_week.get(week_no, []),
            schedule_for_week=schedule.get(week_no, {}),
            holidays=holidays,
            hub_name=hub_name,
            logo_text=logo_text,
            notes=notes,
        )
        for week_no, week_dates in weeks.items()
    ]
    return f"<!doctype html><html lang='id'><head><meta charset='utf-8'>{build_schedule_css()}</head><body>{''.join(pages)}</body></html>"


# -----------------------------------------------------------------------------
# EXCEL
# -----------------------------------------------------------------------------

def build_excel_file(
    assignments_df: pd.DataFrame,
    agents: Sequence[AgentRecord],
    weeks: Dict[int, List[date]],
    holidays: Dict[date, HolidayInfo],
    slots_by_week: Dict[int, List[ShiftSlot]],
    schedule: Dict[int, Dict[str, List[AssignmentEntry]]],
    hub_name: str,
    target_year: int,
    target_month: int,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Overview"

    yellow = "FFD10A"
    black = "111111"
    gray = "F3F3F3"
    red = "D92D20"
    white = "FFFFFF"
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:H1")
    ws["A1"] = f"FLOOR TIME SCHEDULE - {hub_name}"
    ws["A1"].font = Font(bold=True, size=16, color=black)
    ws["A1"].fill = PatternFill("solid", fgColor=yellow)
    ws["A1"].alignment = Alignment(horizontal="center")

    py, pm = previous_month(target_year, target_month)
    overview = [
        ("Target Schedule", month_label(target_year, target_month)),
        ("Data Reference", month_label(py, pm)),
        ("Total Agen Bertugas", len(agents)),
        ("Hari Libur", len(holidays)),
        ("Total Assignment", len(assignments_df)),
    ]
    for idx, (label, value) in enumerate(overview, start=3):
        ws.cell(idx, 1, label).font = Font(bold=True)
        ws.cell(idx, 2, value)
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 28

    # Data agen bulan sebelumnya.
    row_start = 11
    headers = ["Agent ID", "Nama", "Kode", "Jabatan", "Unit", "Total Kehadiran"]
    for col_idx, header in enumerate(headers, 1):
        c = ws.cell(row_start, col_idx, header)
        c.font = Font(bold=True, color=white)
        c.fill = PatternFill("solid", fgColor=black)
        c.border = border
    for r, agent in enumerate(agents, row_start + 1):
        vals = [
            agent.agent_id,
            agent.name,
            agent.code,
            agent.job_title,
            agent.unit,
            agent.previous_month_attendance,
        ]
        for cidx, val in enumerate(vals, 1):
            cell = ws.cell(r, cidx, val)
            cell.border = border
    widths = [18, 30, 14, 24, 38, 16]
    for idx, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    # Sheet per periode (nama internal saja; poster tidak menampilkan nomor minggu).
    for week_no, dates_in_week in weeks.items():
        ws_week = wb.create_sheet(title=f"Periode {week_no}"[:31])
        ws_week.merge_cells("A1:E1")
        ws_week["A1"] = "FLOOR TIME SCHEDULE"
        ws_week["A1"].font = Font(bold=True, size=15)
        ws_week["A1"].fill = PatternFill("solid", fgColor=yellow)
        ws_week["A1"].alignment = Alignment(horizontal="center")
        ws_week.merge_cells("A2:E2")
        ws_week["A2"] = week_range_label(dates_in_week)
        ws_week["A2"].alignment = Alignment(horizontal="center")
        ws_week["A2"].font = Font(bold=True)

        headers_week = ["Hari", "Tanggal", "Shift / Keterangan", "Jam", "Agen"]
        for cidx, header in enumerate(headers_week, 1):
            c = ws_week.cell(4, cidx, header)
            c.font = Font(bold=True, color=white)
            c.fill = PatternFill("solid", fgColor=black)
            c.border = border

        row_idx = 5
        slots_for_date: Dict[date, List[ShiftSlot]] = {}
        for slot in slots_by_week.get(week_no, []):
            slots_for_date.setdefault(slot.tanggal, []).append(slot)

        for d in dates_in_week:
            if d in holidays:
                vals = [DAY_NAMES_ID[d.weekday()], format_date_id(d), holidays[d].name, "-", "-"]
                for cidx, val in enumerate(vals, 1):
                    c = ws_week.cell(row_idx, cidx, val)
                    c.border = border
                    c.font = Font(color=red, bold=(cidx in (1, 3)))
                    c.fill = PatternFill("solid", fgColor="FFF7F6")
                row_idx += 1
                continue
            for slot in sorted(slots_for_date.get(d, []), key=lambda s: s.shift_index):
                entries = schedule.get(week_no, {}).get(slot.key, [])
                agent_text = "\n".join(f"• {e.display_name}" for e in entries) or "Slot Belum Terisi"
                vals = [slot.day_name, format_date_id(d), slot.shift_name, slot.time_label, agent_text]
                for cidx, val in enumerate(vals, 1):
                    c = ws_week.cell(row_idx, cidx, val)
                    c.border = border
                    c.alignment = Alignment(vertical="top", wrap_text=True)
                ws_week.row_dimensions[row_idx].height = max(22, 18 * max(1, len(entries)))
                row_idx += 1
        for idx, w in enumerate([16, 19, 30, 18, 42], 1):
            ws_week.column_dimensions[get_column_letter(idx)].width = w

    ws_detail = wb.create_sheet(title="Data Detail")
    detail = assignments_df.drop(columns=["Urgent"], errors="ignore")
    if detail.empty:
        ws_detail["A1"] = "Belum ada data."
    else:
        for cidx, col in enumerate(detail.columns, 1):
            c = ws_detail.cell(1, cidx, col)
            c.font = Font(bold=True)
            c.fill = PatternFill("solid", fgColor=gray)
            c.border = border
        for ridx, row in enumerate(detail.itertuples(index=False), 2):
            for cidx, val in enumerate(row, 1):
                c = ws_detail.cell(ridx, cidx, val)
                c.border = border
        for cidx in range(1, len(detail.columns) + 1):
            ws_detail.column_dimensions[get_column_letter(cidx)].width = 22

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# -----------------------------------------------------------------------------
# RENDER PNG/PDF — DESAIN BARU, TINGGI DINAMIS
# -----------------------------------------------------------------------------

_FONT_CACHE: Dict[Tuple[int, bool], ImageFont.ImageFont] = {}
_FONT_SOURCE_CACHE: Dict[bool, str] = {}


def _font_candidates(bold: bool) -> List[str]:
    """Return cross-platform corporate sans-serif font candidates.

    The previous renderer only knew Linux font paths. On Windows that made
    Pillow fall back to its tiny bitmap font, so exported posters looked as if
    all text had been shrunk. These candidates intentionally use fonts that are
    commonly present on Windows/macOS/Linux without bundling any font files.
    """
    windows_dir = os.environ.get("WINDIR", r"C:\Windows")
    local_app_data = os.environ.get("LOCALAPPDATA", "")
    windows_fonts = os.path.join(windows_dir, "Fonts")
    user_fonts = os.path.join(local_app_data, "Microsoft", "Windows", "Fonts") if local_app_data else ""

    regular_names = [
        "Aptos.ttf", "AptosDisplay.ttf", "segoeui.ttf", "arial.ttf",
        "calibri.ttf", "DejaVuSans.ttf", "LiberationSans-Regular.ttf",
    ]
    bold_names = [
        "Aptos-Bold.ttf", "AptosDisplay-Bold.ttf", "segoeuib.ttf", "arialbd.ttf",
        "calibrib.ttf", "DejaVuSans-Bold.ttf", "LiberationSans-Bold.ttf",
    ]
    names = bold_names if bold else regular_names

    candidates: List[str] = []
    for folder in [windows_fonts, user_fonts]:
        if folder:
            candidates.extend(os.path.join(folder, name) for name in names)

    # macOS system fonts
    candidates.extend([
        "/System/Library/Fonts/Supplemental/Arial Bold.ttf" if bold else "/System/Library/Fonts/Supplemental/Arial.ttf",
        "/System/Library/Fonts/Supplemental/Helvetica.ttc",
    ])

    # Linux fonts used by the development/test environment.
    candidates.extend([
        "/usr/share/fonts/opentype/inter/Inter-Bold.otf" if bold else "/usr/share/fonts/opentype/inter/Inter-Regular.otf",
        "/usr/share/fonts/opentype/inter/InterDisplay-Bold.otf" if bold else "/usr/share/fonts/opentype/inter/InterDisplay-Regular.otf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf" if bold else "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/truetype/liberation2/LiberationSans-Bold.ttf" if bold else "/usr/share/fonts/truetype/liberation2/LiberationSans-Regular.ttf",
    ])

    # Pillow/FreeType can often resolve these by filename even when the absolute
    # path differs, so keep them as a final scalable-font fallback.
    candidates.extend(names)
    return candidates


def find_inter_font(size: int, bold: bool = False) -> ImageFont.ImageFont:
    key = (size, bold)
    if key in _FONT_CACHE:
        return _FONT_CACHE[key]

    for candidate in _font_candidates(bold):
        try:
            font = ImageFont.truetype(candidate, size=size)
            _FONT_CACHE[key] = font
            _FONT_SOURCE_CACHE[bold] = candidate
            return font
        except (OSError, ValueError):
            continue

    # Newer Pillow versions support a scalable built-in fallback. Prefer it to
    # the legacy tiny bitmap font. This branch should rarely be reached on a
    # normal Windows installation because Segoe UI/Arial are available.
    try:
        font = ImageFont.load_default(size=size)
    except TypeError:
        font = ImageFont.load_default()
    _FONT_CACHE[key] = font
    _FONT_SOURCE_CACHE[bold] = "Pillow built-in fallback"
    return font


def resolved_font_source(bold: bool = False) -> str:
    """Expose the active font source for diagnostics in the UI/tests."""
    find_inter_font(16, bold=bold)
    return _FONT_SOURCE_CACHE.get(bold, "Unknown")


def text_size(draw: ImageDraw.ImageDraw, text: str, font: ImageFont.ImageFont) -> Tuple[int, int]:
    bbox = draw.textbbox((0, 0), text, font=font)
    return bbox[2] - bbox[0], bbox[3] - bbox[1]


def draw_clock_icon(draw: ImageDraw.ImageDraw, center: Tuple[int, int], radius: int, color: str, width: int) -> None:
    cx, cy = center
    draw.ellipse((cx - radius, cy - radius, cx + radius, cy + radius), outline=color, width=width)
    draw.line((cx, cy, cx, cy - radius + width * 2), fill=color, width=width)
    draw.line((cx, cy, cx + radius - width * 2, cy + width), fill=color, width=width)


def draw_brighton_logo(draw: ImageDraw.ImageDraw, center_x: int, y: int, font: ImageFont.ImageFont, logo_text: str) -> None:
    logo = clean_text(logo_text) or DEFAULT_LOGO_TEXT
    if logo.casefold() != "brighton":
        w, _ = text_size(draw, logo, font)
        draw.text((center_x - w // 2, y), logo, font=font, fill=BRIGHTON_BLACK)
        return
    parts = [("Bright", BRIGHTON_BLACK), ("o", BRIGHTON_YELLOW), ("n", BRIGHTON_BLACK)]
    widths = [text_size(draw, p, font)[0] for p, _ in parts]
    x = center_x - sum(widths) // 2
    for (part, color), w in zip(parts, widths):
        draw.text((x, y), part, font=font, fill=color)
        x += w


def rounded_rect(draw: ImageDraw.ImageDraw, box: Tuple[int, int, int, int], radius: int, fill: str, outline: str, width: int) -> None:
    draw.rounded_rectangle(box, radius=radius, fill=fill, outline=outline, width=width)


def split_entries_columns(entries: Sequence[AssignmentEntry], threshold: int = 12) -> List[List[AssignmentEntry]]:
    if len(entries) <= threshold:
        return [list(entries)]
    midpoint = ceil(len(entries) / 2)
    return [list(entries[:midpoint]), list(entries[midpoint:])]


def calculate_week_image_metrics(
    week_dates: Sequence[date],
    slots: Sequence[ShiftSlot],
    schedule_for_week: Dict[str, List[AssignmentEntry]],
    holidays: Dict[date, HolidayInfo],
    orientation: str,
) -> Dict[str, object]:
    """Ukuran font tidak pernah dikecilkan; tinggi canvas mengikuti isi."""
    scale = EXPORT_RENDER_SCALE
    base_width = 1680 if orientation == "Landscape" else 1080
    width = base_width * scale
    margin = (58 if orientation == "Landscape" else 50) * scale
    header_h = (300 if orientation == "Landscape" else 310) * scale
    card_gap = 10 * scale
    notes_h = 116 * scale
    footer_h = 18 * scale
    bottom_pad = 26 * scale
    agent_row_h = (27 if orientation == "Landscape" else 27) * scale
    normal_min_h = (158 if orientation == "Landscape" else 164) * scale
    holiday_h = (108 if orientation == "Landscape" else 116) * scale

    slot_by_date: Dict[date, List[ShiftSlot]] = {}
    for slot in slots:
        slot_by_date.setdefault(slot.tanggal, []).append(slot)

    day_heights: Dict[date, int] = {}
    for d in week_dates:
        if d in holidays:
            day_heights[d] = holiday_h
            continue
        day_slots = slot_by_date.get(d, [])
        max_rows = 1
        for slot in day_slots:
            entries = schedule_for_week.get(slot.key, [])
            columns = split_entries_columns(entries)
            max_rows = max(max_rows, max((len(col) for col in columns), default=1))
        needed = (86 * scale) + max_rows * agent_row_h
        day_heights[d] = max(normal_min_h, needed)

    total_cards = sum(day_heights.values()) + max(0, len(week_dates) - 1) * card_gap
    min_height = (1528 if orientation == "Portrait" else 1188) * scale
    height = max(min_height, header_h + total_cards + notes_h + footer_h + bottom_pad + 28 * scale)
    return {
        "scale": scale,
        "width": width,
        "height": height,
        "margin": margin,
        "header_h": header_h,
        "card_gap": card_gap,
        "notes_h": notes_h,
        "footer_h": footer_h,
        "day_heights": day_heights,
        "agent_row_h": agent_row_h,
    }


def draw_agent_entries(
    draw: ImageDraw.ImageDraw,
    entries: Sequence[AssignmentEntry],
    x: int,
    y: int,
    width: int,
    row_h: int,
    font: ImageFont.ImageFont,
    scale: int,
) -> None:
    columns = split_entries_columns(entries)
    if not entries:
        muted_font = find_inter_font(14 * scale, bold=False)
        draw.text((x, y), "Slot Belum Terisi", font=muted_font, fill="#8A8A8A")
        return

    col_gap = 16 * scale
    col_w = (width - col_gap * (len(columns) - 1)) // len(columns)
    bullet_r = 3 * scale
    for col_idx, col in enumerate(columns):
        cx = x + col_idx * (col_w + col_gap)
        for row_idx, entry in enumerate(col):
            ry = y + row_idx * row_h
            draw.ellipse((cx, ry + 8 * scale, cx + bullet_r * 2, ry + 14 * scale), fill=BRIGHTON_BLACK)
            text_x = cx + 16 * scale
            # Tidak mengecilkan font. Bila nama panjang, wrap maksimal dua baris.
            max_text_w = col_w - 18 * scale
            text = entry.display_name
            if text_size(draw, text, font)[0] <= max_text_w:
                draw.text((text_x, ry), text, font=font, fill=BRIGHTON_BLACK)
            else:
                words = text.split()
                line1 = ""
                line2 = ""
                for word in words:
                    candidate = (line1 + " " + word).strip()
                    if not line2 and text_size(draw, candidate, font)[0] <= max_text_w:
                        line1 = candidate
                    else:
                        line2 = (line2 + " " + word).strip()
                draw.text((text_x, ry - 2 * scale), line1, font=font, fill=BRIGHTON_BLACK)
                if line2:
                    small = find_inter_font(12 * scale, bold=False)
                    draw.text((text_x, ry + 15 * scale), line2, font=small, fill=TEXT_SECONDARY)


def render_week_image(
    week_dates: Sequence[date],
    slots: Sequence[ShiftSlot],
    schedule_for_week: Dict[str, List[AssignmentEntry]],
    holidays: Dict[date, HolidayInfo],
    hub_name: str,
    logo_text: str,
    notes: Sequence[str],
    orientation: str,
) -> Image.Image:
    metrics = calculate_week_image_metrics(week_dates, slots, schedule_for_week, holidays, orientation)
    scale = int(metrics["scale"])
    width = int(metrics["width"])
    height = int(metrics["height"])
    margin = int(metrics["margin"])
    card_gap = int(metrics["card_gap"])
    notes_h = int(metrics["notes_h"])
    footer_h = int(metrics["footer_h"])
    day_heights: Dict[date, int] = metrics["day_heights"]  # type: ignore[assignment]
    agent_row_h = int(metrics["agent_row_h"])

    img = Image.new("RGB", (width, height), "#FFFFFF")
    draw = ImageDraw.Draw(img)

    # Header modern: putih, logo tengah, sudut kuning minimal seperti referensi.
    corner = 110 * scale
    draw.polygon([(0, 0), (corner, 0), (34 * scale, 118 * scale), (0, 118 * scale)], fill=BRIGHTON_YELLOW)
    draw.polygon([(width, 0), (width - corner, 0), (width - 34 * scale, 118 * scale), (width, 118 * scale)], fill=BRIGHTON_YELLOW)
    draw.line((72 * scale, 16 * scale, 20 * scale, 90 * scale), fill=BRIGHTON_YELLOW, width=2 * scale)
    draw.line((width - 72 * scale, 16 * scale, width - 20 * scale, 90 * scale), fill=BRIGHTON_YELLOW, width=2 * scale)

    logo_font = find_inter_font(58 * scale, bold=True)
    draw_brighton_logo(draw, width // 2, 22 * scale, logo_font, logo_text)
    hub_font = find_inter_font(16 * scale, bold=True)
    hub = clean_text(hub_name) or DEFAULT_HUB_NAME
    hub_w, _ = text_size(draw, hub, hub_font)
    hub_y = 102 * scale
    line_y = hub_y + 9 * scale
    draw.line((width // 2 - 170 * scale, line_y, width // 2 - 95 * scale, line_y), fill=BRIGHTON_YELLOW, width=2 * scale)
    draw.text((width // 2 - hub_w // 2, hub_y), hub, font=hub_font, fill=BRIGHTON_BLACK)
    draw.line((width // 2 + 95 * scale, line_y, width // 2 + 170 * scale, line_y), fill=BRIGHTON_YELLOW, width=2 * scale)

    title_font = find_inter_font(39 * scale, bold=True)
    title = "FLOOR TIME SCHEDULE"
    tw, th = text_size(draw, title, title_font)
    title_y = 150 * scale
    draw.text((width // 2 - tw // 2, title_y), title, font=title_font, fill=BRIGHTON_BLACK)

    range_font = find_inter_font(20 * scale, bold=True)
    range_text = week_range_label(week_dates)
    rw, _ = text_size(draw, range_text, range_font)
    draw.text((width // 2 - rw // 2, title_y + th + 24 * scale), range_text, font=range_font, fill=BRIGHTON_BLACK)

    # Cards.
    slot_by_date: Dict[date, List[ShiftSlot]] = {}
    for slot in slots:
        slot_by_date.setdefault(slot.tanggal, []).append(slot)

    y = 282 * scale
    card_x1 = margin
    card_x2 = width - margin
    card_w = card_x2 - card_x1
    day_col_w = int(card_w * 0.25)
    shift_col_w = (card_w - day_col_w) // 2

    day_font = find_inter_font(27 * scale, bold=True)
    date_font = find_inter_font(16 * scale, bold=False)
    time_font = find_inter_font(16 * scale, bold=True)
    agent_font = find_inter_font(14 * scale, bold=False)
    holiday_font = find_inter_font(18 * scale, bold=True)
    badge_font = find_inter_font(10 * scale, bold=True)

    for d in week_dates:
        card_h = day_heights[d]
        is_holiday = d in holidays
        fill = HOLIDAY_BG if is_holiday else "#FFFFFF"
        outline = "#ED9A94" if is_holiday else BORDER_COLOR
        rounded_rect(draw, (card_x1, y, card_x2, y + card_h), 10 * scale, fill, outline, 1 * scale)

        divider_x = card_x1 + day_col_w
        divider_color = "#ED9A94" if is_holiday else "#ECECEC"
        draw.line((divider_x, y + 16 * scale, divider_x, y + card_h - 16 * scale), fill=divider_color, width=1 * scale)

        meta_x = card_x1 + 38 * scale
        day_color = HOLIDAY_RED if is_holiday else BRIGHTON_BLACK
        draw.text((meta_x, y + 28 * scale), DAY_NAMES_ID[d.weekday()], font=day_font, fill=day_color)
        draw.text((meta_x, y + 68 * scale), format_date_id(d), font=date_font, fill=day_color)
        accent_color = HOLIDAY_RED if is_holiday else BRIGHTON_YELLOW
        draw.rectangle((meta_x, y + 104 * scale, meta_x + 44 * scale, y + 106 * scale), fill=accent_color)

        if is_holiday:
            name_x = divider_x + 44 * scale
            name_y = y + (card_h - text_size(draw, holidays[d].name, holiday_font)[1]) // 2 - 2 * scale
            draw.text((name_x, name_y), holidays[d].name, font=holiday_font, fill=HOLIDAY_RED)
            y += card_h + card_gap
            continue

        day_slots = sorted(slot_by_date.get(d, []), key=lambda s: s.shift_index)
        for idx, slot in enumerate(day_slots[:2]):
            sx = divider_x + idx * shift_col_w
            if idx == 1:
                draw.line((sx, y + 22 * scale, sx, y + card_h - 22 * scale), fill="#D8D8D8", width=1 * scale)
            pad = 30 * scale
            content_x = sx + pad
            clock_center = (content_x + 14 * scale, y + 38 * scale)
            draw_clock_icon(draw, clock_center, 12 * scale, BRIGHTON_YELLOW, 2 * scale)
            draw.text((content_x + 38 * scale, y + 25 * scale), slot.time_label, font=time_font, fill=BRIGHTON_BLACK)
            draw.line((content_x + 38 * scale, y + 52 * scale, content_x + 175 * scale, y + 52 * scale), fill=BRIGHTON_YELLOW, width=1 * scale)

            if d.weekday() == 5 and idx == 1:
                badge = "Jam Operasional Khusus"
                bw, bh = text_size(draw, badge, badge_font)
                bx2 = sx + shift_col_w - 16 * scale
                bx1 = bx2 - bw - 18 * scale
                by1 = y + 8 * scale
                rounded_rect(draw, (bx1, by1, bx2, by1 + bh + 10 * scale), 5 * scale, BRIGHTON_YELLOW, BRIGHTON_YELLOW, 1)
                draw.text((bx1 + 9 * scale, by1 + 5 * scale), badge, font=badge_font, fill=BRIGHTON_BLACK)

            entries = schedule_for_week.get(slot.key, [])
            draw_agent_entries(
                draw,
                entries,
                x=content_x,
                y=y + 70 * scale,
                width=shift_col_w - 2 * pad,
                row_h=agent_row_h,
                font=agent_font,
                scale=scale,
            )

        y += card_h + card_gap

    # Catatan sederhana tanpa icon besar.
    notes_y = y + 4 * scale
    notes_w = card_w
    rounded_rect(draw, (card_x1, notes_y, card_x2, notes_y + notes_h), 10 * scale, "#FAFAFA", BORDER_COLOR, 1 * scale)
    note_title_font = find_inter_font(18 * scale, bold=True)
    note_font = find_inter_font(13 * scale, bold=False)
    title_x = card_x1 + 34 * scale
    draw.text((title_x, notes_y + 30 * scale), "CATATAN", font=note_title_font, fill=BRIGHTON_BLACK)
    sep_x = title_x + 120 * scale
    draw.rectangle((sep_x, notes_y + 24 * scale, sep_x + 3 * scale, notes_y + notes_h - 24 * scale), fill=BRIGHTON_YELLOW)
    note_x = sep_x + 34 * scale
    note_y = notes_y + 24 * scale
    clean_notes = [clean_text(n) for n in notes if clean_text(n)] or ["Agen yang mendapatkan jadwal floor time masih berada di kantor."]
    for note in clean_notes[:3]:
        draw.ellipse((note_x, note_y + 7 * scale, note_x + 6 * scale, note_y + 13 * scale), fill=BRIGHTON_BLACK)
        draw.text((note_x + 18 * scale, note_y), note, font=note_font, fill=BRIGHTON_BLACK)
        note_y += 28 * scale

    # Footer accent.
    draw.rectangle((0, height - footer_h, width, height), fill=BRIGHTON_YELLOW)
    draw.polygon([
        (int(width * 0.80), height - footer_h),
        (int(width * 0.86), height - footer_h),
        (int(width * 0.84), height),
        (int(width * 0.78), height),
    ], fill="#151515")
    return img


def render_all_week_images(
    weeks: Dict[int, List[date]],
    slots_by_week: Dict[int, List[ShiftSlot]],
    schedule: Dict[int, Dict[str, List[AssignmentEntry]]],
    holidays: Dict[date, HolidayInfo],
    hub_name: str,
    logo_text: str,
    notes: Sequence[str],
    orientation: str,
) -> Dict[int, Image.Image]:
    return {
        week_no: render_week_image(
            week_dates=week_dates,
            slots=slots_by_week.get(week_no, []),
            schedule_for_week=schedule.get(week_no, {}),
            holidays=holidays,
            hub_name=hub_name,
            logo_text=logo_text,
            notes=notes,
            orientation=orientation,
        )
        for week_no, week_dates in weeks.items()
    }


ROSTER_COLUMN_OPTIONS = ("Nama", "Jabatan", "Unit", "Total Kehadiran")


def _roster_cell_value(agent: AgentRecord, column: str) -> str:
    if column == "Nama":
        return agent.display_name
    if column == "Jabatan":
        return agent.job_title or "-"
    if column == "Unit":
        return agent.unit or "-"
    if column == "Total Kehadiran":
        return str(agent.previous_month_attendance)
    return ""


def _wrap_text_for_width(
    draw: ImageDraw.ImageDraw,
    text: str,
    font: ImageFont.ImageFont,
    max_width: int,
) -> List[str]:
    """Wrap teks tanpa mengecilkan font; tinggi row mengikuti jumlah baris."""
    value = clean_text(text) or "-"
    words = value.split()
    if not words:
        return ["-"]
    lines: List[str] = []
    current = ""
    for word in words:
        candidate = f"{current} {word}".strip()
        if not current or text_size(draw, candidate, font)[0] <= max_width:
            current = candidate
            continue
        lines.append(current)
        current = word
    if current:
        lines.append(current)
    return lines or [value]


def render_monthly_agent_roster_image(
    agents: Sequence[AgentRecord],
    target_year: int,
    target_month: int,
    selected_columns: Sequence[str],
    sort_mode: str,
    hub_name: str,
    logo_text: str,
) -> Image.Image:
    """Render daftar agen bulanan dengan visual yang selaras dengan poster Floor Time.

    Semua agen aktif ditampilkan. Font tidak diperkecil berdasarkan jumlah agen;
    canvas dan tinggi setiap baris bertambah mengikuti konten.
    """
    columns = [c for c in ROSTER_COLUMN_OPTIONS if c in set(selected_columns)]
    if not columns:
        columns = ["Nama"]
    ordered_agents = sort_agents_for_roster(agents, sort_mode)

    scale = 1
    width = 1080
    margin = 50
    top_table = 258
    header_row_h = 48
    footer_h = 18
    bottom_pad = 34
    table_w = width - 2 * margin

    weight_map = {
        "Nama": 0.32,
        "Jabatan": 0.28,
        "Unit": 0.20,
        "Total Kehadiran": 0.20,
    }
    total_weight = sum(weight_map[c] for c in columns)
    col_widths = [int(table_w * weight_map[c] / total_weight) for c in columns]
    if col_widths:
        col_widths[-1] += table_w - sum(col_widths)

    body_font = find_inter_font(14 * scale, bold=False)
    header_font = find_inter_font(14 * scale, bold=True)
    dummy = Image.new("RGB", (width, 200), "#FFFFFF")
    dummy_draw = ImageDraw.Draw(dummy)

    row_layouts: List[Tuple[int, List[List[str]]]] = []
    for agent in ordered_agents:
        wrapped_cells: List[List[str]] = []
        max_lines = 1
        for col, col_w in zip(columns, col_widths):
            # Total Kehadiran tidak perlu wrap dan rata kanan saat digambar.
            lines = [_roster_cell_value(agent, col)] if col == "Total Kehadiran" else _wrap_text_for_width(
                dummy_draw,
                _roster_cell_value(agent, col),
                body_font,
                max(40, col_w - 24),
            )
            wrapped_cells.append(lines)
            max_lines = max(max_lines, len(lines))
        row_h = max(44, 18 * max_lines + 18)
        row_layouts.append((row_h, wrapped_cells))

    rows_h = sum(row_h for row_h, _ in row_layouts)
    height = max(820, top_table + header_row_h + rows_h + bottom_pad + footer_h)
    img = Image.new("RGB", (width, height), "#FFFFFF")
    draw = ImageDraw.Draw(img)

    # Header mengikuti bahasa visual poster mingguan.
    corner = 110
    draw.polygon([(0, 0), (corner, 0), (34, 118), (0, 118)], fill=BRIGHTON_YELLOW)
    draw.polygon([(width, 0), (width - corner, 0), (width - 34, 118), (width, 118)], fill=BRIGHTON_YELLOW)
    draw.line((72, 16, 20, 90), fill=BRIGHTON_YELLOW, width=2)
    draw.line((width - 72, 16, width - 20, 90), fill=BRIGHTON_YELLOW, width=2)

    logo_font = find_inter_font(58, bold=True)
    draw_brighton_logo(draw, width // 2, 22, logo_font, logo_text)
    hub_font = find_inter_font(16, bold=True)
    hub = clean_text(hub_name) or DEFAULT_HUB_NAME
    hub_w, _ = text_size(draw, hub, hub_font)
    hub_y = 102
    line_y = hub_y + 9
    draw.line((width // 2 - 170, line_y, width // 2 - 95, line_y), fill=BRIGHTON_YELLOW, width=2)
    draw.text((width // 2 - hub_w // 2, hub_y), hub, font=hub_font, fill=BRIGHTON_BLACK)
    draw.line((width // 2 + 95, line_y, width // 2 + 170, line_y), fill=BRIGHTON_YELLOW, width=2)

    title_font = find_inter_font(39, bold=True)
    title = f"FLOOR TIME {MONTH_NAMES_ID[target_month].upper()}"
    tw, th = text_size(draw, title, title_font)
    title_y = 156
    draw.text((width // 2 - tw // 2, title_y), title, font=title_font, fill=BRIGHTON_BLACK)

    year_font = find_inter_font(21, bold=True)
    year_text = str(target_year)
    yw, yh = text_size(draw, year_text, year_font)
    draw.text((width // 2 - yw // 2, title_y + th + 14), year_text, font=year_font, fill=TEXT_SECONDARY)

    # Table container + header.
    table_x1 = margin
    table_x2 = width - margin
    table_y1 = top_table
    table_y2 = top_table + header_row_h + rows_h
    rounded_rect(draw, (table_x1, table_y1, table_x2, table_y2), 10, "#FFFFFF", BORDER_COLOR, 1)
    draw.rounded_rectangle((table_x1, table_y1, table_x2, table_y1 + header_row_h + 8), radius=10, fill=BRIGHTON_BLACK)
    draw.rectangle((table_x1, table_y1 + header_row_h - 8, table_x2, table_y1 + header_row_h + 1), fill=BRIGHTON_BLACK)
    draw.rectangle((table_x1, table_y1, table_x2, table_y1 + 3), fill=BRIGHTON_YELLOW)

    x_positions = [table_x1]
    x = table_x1
    for w in col_widths:
        x += w
        x_positions.append(x)

    for idx, (column, x1, x2) in enumerate(zip(columns, x_positions[:-1], x_positions[1:])):
        label = column
        if column == "Total Kehadiran":
            label = "Total Kehadiran"
        lw, lh = text_size(draw, label, header_font)
        if column == "Total Kehadiran":
            tx = x2 - 14 - lw
        else:
            tx = x1 + 14
        draw.text((tx, table_y1 + (header_row_h - lh) // 2 - 1), label, font=header_font, fill="#FFFFFF")
        if idx > 0:
            draw.line((x1, table_y1 + 10, x1, table_y2 - 1), fill="#E2E2E2", width=1)

    y = table_y1 + header_row_h
    line_h = 18
    for row_index, (row_h, wrapped_cells) in enumerate(row_layouts):
        row_fill = "#FFFFFF" if row_index % 2 == 0 else "#F1F1F1"
        draw.rectangle((table_x1 + 1, y, table_x2 - 1, y + row_h), fill=row_fill)
        draw.line((table_x1, y, table_x2, y), fill="#E7E7E7", width=1)

        for column, x1, x2, lines in zip(columns, x_positions[:-1], x_positions[1:], wrapped_cells):
            if column == "Total Kehadiran":
                value = lines[0]
                vw, vh = text_size(draw, value, body_font)
                draw.text((x2 - 14 - vw, y + (row_h - vh) // 2 - 1), value, font=body_font, fill=BRIGHTON_BLACK)
                continue
            text_block_h = len(lines) * line_h
            ty = y + max(8, (row_h - text_block_h) // 2)
            for line in lines:
                draw.text((x1 + 14, ty), line, font=body_font, fill=BRIGHTON_BLACK)
                ty += line_h
        y += row_h

    # Footer accent sama seperti poster mingguan.
    draw.rectangle((0, height - footer_h, width, height), fill=BRIGHTON_YELLOW)
    draw.polygon([
        (int(width * 0.80), height - footer_h),
        (int(width * 0.86), height - footer_h),
        (int(width * 0.84), height),
        (int(width * 0.78), height),
    ], fill="#151515")
    return img


def build_images_zip(
    week_images: Dict[int, Image.Image],
    weeks: Dict[int, List[date]],
    year: int,
    month: int,
    monthly_roster_image: Optional[Image.Image] = None,
) -> bytes:
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        if monthly_roster_image is not None:
            roster_buffer = io.BytesIO()
            monthly_roster_image.save(roster_buffer, format="PNG", dpi=(300, 300), compress_level=2)
            zf.writestr(f"floor_time_{year}-{month:02d}_bulanan.png", roster_buffer.getvalue())
        for week_no, img in week_images.items():
            img_buffer = io.BytesIO()
            img.save(img_buffer, format="PNG", dpi=(300, 300), compress_level=2)
            dates = weeks.get(week_no, [])
            if dates:
                start, end = min(dates), max(dates)
                filename = f"floor_time_{start:%Y%m%d}_{end:%Y%m%d}.png"
            else:
                filename = f"floor_time_{year}-{month:02d}_periode_{week_no}.png"
            zf.writestr(filename, img_buffer.getvalue())
    return buffer.getvalue()


def build_pdf_from_images(week_images: Dict[int, Image.Image], monthly_roster_image: Optional[Image.Image] = None) -> bytes:
    pages: List[Image.Image] = []
    if monthly_roster_image is not None:
        pages.append(monthly_roster_image.convert("RGB"))
    pages.extend(img.convert("RGB") for _, img in sorted(week_images.items()))
    if not pages:
        return b""
    buffer = io.BytesIO()
    first, rest = pages[0], pages[1:]
    first.save(buffer, format="PDF", save_all=True, append_images=rest, resolution=300.0)
    return buffer.getvalue()


# -----------------------------------------------------------------------------
# STREAMLIT UI
# -----------------------------------------------------------------------------

def inject_app_style() -> None:
    """Force a stable Brighton light UI independent of the user's Streamlit theme."""
    st.markdown(
        f"""
        <style>
        :root {{
            --brighton-yellow:{BRIGHTON_YELLOW};
            --brighton-black:{BRIGHTON_BLACK};
            --holiday-red:{HOLIDAY_RED};
            --surface:#ffffff;
            --surface-soft:#f7f7f7;
            --border:#e4e4e4;
            --muted:#666666;
        }}

        /* The app used to force a white background while Streamlit could still
           be in dark mode. That produced white/light-gray widget text on white. */
        html, body, [data-testid="stAppViewContainer"], .stApp {{
            background:var(--surface) !important;
            color:var(--brighton-black) !important;
        }}
        [data-testid="stHeader"] {{ background:rgba(255,255,255,.96) !important; }}
        [data-testid="stSidebar"] {{ background:#fafafa !important; border-right:1px solid #ececec; }}
        [data-testid="stSidebar"] > div:first-child {{ background:#fafafa !important; }}

        .block-container {{ max-width: 1380px; padding-top: 1.35rem; padding-bottom: 3rem; }}

        /* Keep all native Streamlit text legible even if browser preference is dark. */
        .stMarkdown, .stMarkdown p, .stMarkdown li, .stMarkdown h1, .stMarkdown h2,
        .stMarkdown h3, .stMarkdown h4, .stCaption, [data-testid="stCaptionContainer"],
        [data-testid="stWidgetLabel"] p, [data-testid="stMetricLabel"],
        [data-testid="stMetricValue"], [data-testid="stMetricDelta"],
        [data-testid="stExpander"] summary, [data-testid="stFileUploader"] span,
        [data-testid="stFileUploader"] small {{ color:var(--brighton-black) !important; }}

        /* Inputs/selects/textareas */
        input, textarea {{ color:var(--brighton-black) !important; }}
        div[data-baseweb="input"] > div, div[data-baseweb="textarea"] > div,
        div[data-baseweb="select"] > div {{
            background:#ffffff !important;
            color:var(--brighton-black) !important;
            border-color:#d8d8d8 !important;
        }}
        div[data-baseweb="select"] span {{ color:var(--brighton-black) !important; }}

        /* Tabs: strong contrast, Brighton yellow active state. */
        [data-testid="stTabs"] [role="tablist"] {{ gap:1.6rem; border-bottom:1px solid #ececec; }}
        [data-testid="stTabs"] button[role="tab"] {{
            color:#6a6a6a !important;
            font-weight:650 !important;
        }}
        [data-testid="stTabs"] button[role="tab"][aria-selected="true"] {{
            color:var(--brighton-black) !important;
            font-weight:800 !important;
        }}
        [data-testid="stTabs"] [data-baseweb="tab-highlight"] {{ background:var(--brighton-yellow) !important; }}

        /* Metrics and cards */
        [data-testid="stMetric"] {{
            background:#ffffff;
            border:1px solid var(--border);
            border-radius:12px;
            padding:14px 16px;
        }}
        .app-hero {{ border-bottom:3px solid var(--brighton-yellow); padding:0 0 18px; margin-bottom:16px; }}
        .app-hero h1 {{ margin:0; font-family:Inter,Arial,Helvetica,sans-serif; font-size:2.15rem; font-weight:850; color:#111 !important; letter-spacing:-.02em; }}
        .app-hero h1 .o {{ color:var(--brighton-yellow) !important; }}
        .app-hero p {{ margin:.45rem 0 0; color:#5b5b5b !important; font-size:.98rem; }}
        .overview-card {{ border:1px solid var(--border); border-radius:12px; padding:16px 18px; background:#fff; }}
        .reference-label {{ font-size:.82rem; color:#777 !important; text-transform:uppercase; letter-spacing:.08em; }}
        .reference-value {{ font-size:1.2rem; font-weight:750; color:#111 !important; margin-top:4px; }}
        .app-credit {{ text-align:center; color:#999 !important; font-size:.78rem; margin-top:24px; }}

        /* Buttons */
        .stButton > button[kind="primary"], .stDownloadButton > button {{
            border-radius:8px !important;
            font-weight:750 !important;
        }}
        .stButton > button[kind="primary"] {{
            background:var(--brighton-yellow) !important;
            border-color:var(--brighton-yellow) !important;
            color:#111 !important;
        }}
        .stButton > button[kind="primary"] * {{ color:#111 !important; }}

        /* Dataframes/editors get a visible boundary instead of blending into page. */
        [data-testid="stDataFrame"], [data-testid="stDataEditor"] {{
            border:1px solid var(--border);
            border-radius:10px;
            overflow:hidden;
        }}
        </style>
        """,
        unsafe_allow_html=True,
    )


def initialize_session_state() -> None:
    defaults = {
        "assignments_df": pd.DataFrame(),
        "schedule": {},
        "capacity_by_week": {},
        "warnings": [],
        "validation_messages": [],
        "validation_ok": False,
        "excel_bytes": b"",
        "image_zip_bytes": b"",
        "pdf_bytes": b"",
        "week_images": {},
        "monthly_roster_image": None,
        "weeks": {},
        "slots_by_week": {},
        "holidays": {},
        "agents": [],
        "output_orientation": "Portrait",
        "generated_target": None,
        "generated_config": None,
    }
    for key, value in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = value


def clear_generated_outputs() -> None:
    """Hapus output lama ketika data/konfigurasi berubah agar hasil tidak stale."""
    for key, value in {
        "assignments_df": pd.DataFrame(),
        "schedule": {},
        "capacity_by_week": {},
        "warnings": [],
        "validation_messages": [],
        "validation_ok": False,
        "excel_bytes": b"",
        "image_zip_bytes": b"",
        "pdf_bytes": b"",
        "week_images": {},
        "monthly_roster_image": None,
        "slots_by_week": {},
        "holidays": {},
        "generated_target": None,
        "generated_config": None,
    }.items():
        st.session_state[key] = value


def render_credit() -> None:
    st.markdown('<div class="app-credit">created by @rahmathidyatt</div>', unsafe_allow_html=True)


def main() -> None:
    st.set_page_config(page_title=APP_TITLE, page_icon="🟨", layout="wide")
    inject_app_style()
    initialize_session_state()

    st.markdown(
        """
        <div class="app-hero">
          <h1>Bright<span class="o">o</span>n Floor Time Scheduler</h1>
        </div>
        """,
        unsafe_allow_html=True,
    )

    today = datetime.today().date()
    with st.sidebar:
        st.header("Pengaturan Utama")
        hub_name = st.text_input("Nama Hub", value=DEFAULT_HUB_NAME)
        logo_text = st.text_input("Teks Logo", value=DEFAULT_LOGO_TEXT)
        selected_year = int(st.number_input("Tahun Jadwal", min_value=2024, max_value=2100, value=today.year, step=1))
        selected_month = int(st.selectbox("Bulan Jadwal", options=list(MONTH_NAMES_ID), index=today.month - 1, format_func=lambda m: MONTH_NAMES_ID[m]))

        py, pm = previous_month(selected_year, selected_month)
        st.caption(f"Data referensi otomatis: **{month_label(py, pm)}**")

        st.divider()
        st.subheader("Kapasitas & Frekuensi")
        weekday_capacity = int(st.number_input("Target agen per shift Senin-Jumat", min_value=1, max_value=30, value=4, step=1))
        saturday_capacity = int(st.number_input("Target agen per shift Sabtu", min_value=1, max_value=30, value=3, step=1))
        max_weekly_assignments = int(st.radio("Maksimum Floor Time / Agent / Minggu", [1, 2, 3], format_func=lambda x: f"{x}x", horizontal=True))
        auto_expand_capacity = st.checkbox("Perluas kapasitas bila agen lebih banyak", value=True)

        st.divider()
        # Poster aplikasi ini selalu menggunakan orientasi portrait. Opsi orientasi
        # sengaja dihilangkan agar sidebar lebih ringkas dan konsisten dengan output.
        output_orientation = "Portrait"
        use_seed = st.checkbox("Gunakan kode audit", value=True)
        seed_text = st.text_input("Kode audit", value=f"{selected_year}-{selected_month:02d}-brighton") if use_seed else ""

    work_dates = get_month_work_dates(selected_year, selected_month)
    weeks = group_dates_by_calendar_week(work_dates)
    date_labels = [date_option_label(d) for d in work_dates]
    label_to_date = {date_option_label(d): d for d in work_dates}

    tabs = st.tabs(["1. Data & Overview", "2. Konfigurasi & Generate", "3. Jadwal", "4. Download", "5. Panduan"])

    # ------------------------------------------------------------------ DATA
    with tabs[0]:
        uploaded_file = st.file_uploader(
            "Upload file data agen (Excel/CSV)",
            type=["csv", "xlsx", "xls"],
            help="Format utama: No | Nama | Jabatan | Office | Total. Kolom No bersifat opsional.",
        )

        uploaded_agents: List[AgentRecord] = []
        agents: List[AgentRecord] = []
        if uploaded_file is not None:
            try:
                uploaded_df = read_agent_upload(uploaded_file)
                uploaded_agents, upload_warnings = dataframe_to_agents(uploaded_df)
                if not uploaded_agents:
                    st.error("File berhasil dibaca, tetapi tidak ditemukan data agen yang valid pada kolom Nama.")
                elif upload_warnings:
                    st.caption(f"{len(upload_warnings)} baris duplikat/tidak valid diabaikan saat membaca file.")
            except Exception as exc:
                st.error(f"File tidak dapat dibaca: {exc}")

        if uploaded_file is None:
            st.info("Upload file Excel/CSV untuk memulai penyusunan Floor Time.")
        elif uploaded_agents:
            max_attendance = max((a.previous_month_attendance for a in uploaded_agents), default=0)
            filter_col, metric_col = st.columns([0.72, 0.28], gap="large")
            with filter_col:
                minimum_attendance = int(
                    st.number_input(
                        "Minimum Total Kehadiran",
                        min_value=0,
                        max_value=max(31, max_attendance),
                        value=5,
                        step=1,
                        help="Contoh: isi 5 agar agen dengan total kehadiran 1–4 tidak masuk kandidat Floor Time.",
                    )
                )

            attendance_candidates = filter_agents_for_schedule(
                uploaded_agents,
                minimum_attendance=minimum_attendance,
            )

            selector_df = pd.DataFrame([
                {
                    "Masuk Jadwal": True,
                    "Agent ID": a.agent_id,
                    "Nama": a.display_name,
                    "Jabatan": a.job_title or "-",
                    "Unit": a.unit or "-",
                    "Total Kehadiran": a.previous_month_attendance,
                }
                for a in attendance_candidates
            ])

            if selector_df.empty:
                st.warning("Tidak ada agen yang memenuhi minimum Total Kehadiran yang dipilih.")
            else:
                editor_key = f"agent_month_selector_{selected_year}_{selected_month}_{minimum_attendance}_{uploaded_file.name}"
                selected_df = st.data_editor(
                    selector_df,
                    key=editor_key,
                    use_container_width=True,
                    hide_index=True,
                    height=520,
                    disabled=["Agent ID", "Nama", "Jabatan", "Unit", "Total Kehadiran"],
                    column_config={
                        "Masuk Jadwal": st.column_config.CheckboxColumn(
                            "Masuk Jadwal",
                            help="Hilangkan centang untuk mengeluarkan agen dari Floor Time bulan ini.",
                            default=True,
                        ),
                        "Agent ID": None,
                        "Nama": st.column_config.TextColumn("Nama"),
                        "Jabatan": st.column_config.TextColumn("Jabatan"),
                        "Unit": st.column_config.TextColumn("Unit"),
                        "Total Kehadiran": st.column_config.NumberColumn("Total Kehadiran", format="%d"),
                    },
                )

                active_ids = set(
                    selected_df.loc[selected_df["Masuk Jadwal"] == True, "Agent ID"].astype(str).tolist()  # noqa: E712
                )
                excluded_ids = {a.agent_id for a in attendance_candidates if a.agent_id not in active_ids}
                agents = filter_agents_for_schedule(
                    uploaded_agents,
                    minimum_attendance=minimum_attendance,
                    excluded_agent_ids=excluded_ids,
                )

            with metric_col:
                st.metric("Total Agen Bertugas", len(agents))

        st.session_state["current_agents"] = agents
        st.session_state["uploaded_agents"] = uploaded_agents
        render_credit()

    # -------------------------------------------------------- CONFIG & GENERATE
    with tabs[1]:
        agents = st.session_state.get("current_agents", [])
        st.subheader("Konfigurasi Floor Time")
        top_left, top_right = st.columns([1, 1], gap="large")

        with top_left:
            st.markdown("#### Hari Libur / Tanggal Merah")
            selected_holiday_labels = st.multiselect(
                "Pilih tanggal libur",
                options=date_labels,
                help="Tanggal terpilih tidak akan mendapat agent maupun shift.",
            )
            for label in selected_holiday_labels:
                d = label_to_date[label]
                st.text_input(
                    f"Keterangan — {date_option_label(d)}",
                    value="Hari Libur",
                    key=f"holiday_name_{d.isoformat()}",
                    placeholder="Contoh: Maulid Nabi Muhammad SAW",
                )
            holidays = build_holidays(selected_holiday_labels, label_to_date)

        with top_right:
            st.markdown("#### Request Jadwal Khusus Internal")
            agent_options = [a.display_name for a in agents]
            urgent_template = pd.DataFrame(columns=["Agen", "Tanggal", "Shift", "Catatan"])
            urgent_df = st.data_editor(
                urgent_template,
                key="urgent_editor",
                num_rows="dynamic",
                use_container_width=True,
                hide_index=True,
                column_config={
                    "Agen": st.column_config.SelectboxColumn("Agen", options=agent_options or [""]),
                    "Tanggal": st.column_config.SelectboxColumn("Tanggal", options=date_labels or [""]),
                    "Shift": st.column_config.SelectboxColumn("Shift", options=["Pagi", "Siang"]),
                    "Catatan": st.column_config.TextColumn("Catatan"),
                },
            )

        st.markdown("#### Ringkasan Kapasitas")
        slots_preview = build_slots_by_week(weeks, holidays, weekday_capacity, saturday_capacity)
        active_dates = {slot.tanggal for slots in slots_preview.values() for slot in slots}
        m1, m2 = st.columns(2)
        m1.metric("Agent", len(agents))
        m2.metric("Hari Libur", len(holidays))

        # Analisis kapasitas tetap dihitung oleh scheduling engine, tetapi tidak
        # ditampilkan sebagai deretan notifikasi agar halaman tetap bersih.

        notes_default = """Agen yang mendapatkan jadwal floor time masih berada di kantor.
Penanganan klien berdasarkan agen yang melakukan absensi pertama."""
        notes_raw = st.text_area("Catatan pada poster", value=notes_default, height=90)
        notes = [clean_text(line) for line in notes_raw.splitlines() if clean_text(line)]

        current_config = (
            selected_year, selected_month, weekday_capacity, saturday_capacity,
            max_weekly_assignments, auto_expand_capacity, output_orientation,
            seed_text if use_seed else "", hub_name, logo_text,
            tuple((a.agent_id, a.job_title, a.office, a.previous_month_attendance) for a in agents),
            tuple(sorted((d.isoformat(), info.name) for d, info in holidays.items())),
            tuple(notes),
        )
        generated_config = st.session_state.get("generated_config")
        if st.session_state.get("schedule") and generated_config is not None and generated_config != current_config:
            clear_generated_outputs()
            st.info("Data atau konfigurasi berubah. Hasil sebelumnya dinonaktifkan agar tidak menampilkan jadwal yang stale. Klik Generate untuk membuat hasil terbaru.")

        button_label = "Regenerate Jadwal Bulanan" if st.session_state.get("schedule") else "Generate Jadwal Bulanan"
        generate_clicked = st.button(button_label, type="primary", use_container_width=True)

        if generate_clicked:
            if not agents:
                st.error("Daftar agent belum tersedia. Isi data di tab Data & Overview terlebih dahulu.")
            else:
                urgent_requests, urgent_errors = parse_urgent_requests(urgent_df, label_to_date, agents, holidays)
                if urgent_errors:
                    for err in urgent_errors:
                        st.error(err)
                else:
                    slots_by_week = build_slots_by_week(weeks, holidays, weekday_capacity, saturday_capacity)
                    assignments_df, schedule, capacity_by_week, warnings = generate_schedule(
                        agents=agents,
                        slots_by_week=slots_by_week,
                        urgent_requests=urgent_requests,
                        max_weekly_assignments=max_weekly_assignments,
                        auto_expand_capacity=auto_expand_capacity,
                        seed_text=seed_text if use_seed else None,
                    )
                    validation_ok, validation_messages = validate_schedule(
                        assignments_df, agents, weeks, holidays, max_weekly_assignments
                    )

                    week_images = render_all_week_images(
                        weeks=weeks,
                        slots_by_week=slots_by_week,
                        schedule=schedule,
                        holidays=holidays,
                        hub_name=hub_name,
                        logo_text=logo_text,
                        notes=notes,
                        orientation=output_orientation,
                    )
                    monthly_roster_image = render_monthly_agent_roster_image(
                        agents=agents,
                        target_year=selected_year,
                        target_month=selected_month,
                        selected_columns=list(ROSTER_COLUMN_OPTIONS),
                        sort_mode="Default (sesuai file)",
                        hub_name=hub_name,
                        logo_text=logo_text,
                    )
                    excel_bytes = build_excel_file(
                        assignments_df, agents, weeks, holidays, slots_by_week, schedule,
                        hub_name, selected_year, selected_month,
                    )
                    image_zip_bytes = build_images_zip(week_images, weeks, selected_year, selected_month, monthly_roster_image=monthly_roster_image)
                    pdf_bytes = build_pdf_from_images(week_images, monthly_roster_image=monthly_roster_image)

                    st.session_state.update({
                        "assignments_df": assignments_df,
                        "schedule": schedule,
                        "capacity_by_week": capacity_by_week,
                        "warnings": warnings,
                        "validation_messages": validation_messages,
                        "validation_ok": validation_ok,
                        "excel_bytes": excel_bytes,
                        "image_zip_bytes": image_zip_bytes,
                        "pdf_bytes": pdf_bytes,
                        "week_images": week_images,
                        "monthly_roster_image": monthly_roster_image,
                        "weeks": weeks,
                        "slots_by_week": slots_by_week,
                        "holidays": holidays,
                        "agents": agents,
                        "output_orientation": output_orientation,
                        "generated_target": (selected_year, selected_month),
                        "generated_config": current_config,
                    })
                    if validation_ok:
                        st.success("Jadwal berhasil dibuat.")
                    else:
                        st.error("Jadwal terbentuk tetapi masih terdapat pelanggaran validasi. Periksa pesan pada tab Jadwal.")
        render_credit()

    # ----------------------------------------------------------------- RESULT
    with tabs[2]:
        st.subheader("Floor Time Schedule")
        assignments_df = st.session_state.get("assignments_df", pd.DataFrame())
        schedule_state = st.session_state.get("schedule", {})
        weeks_state = st.session_state.get("weeks", {})
        if assignments_df.empty and not schedule_state:
            st.info("Belum ada jadwal. Lakukan Generate pada tab Konfigurasi & Generate.")
        else:
            c1, c2, c3 = st.columns(3)
            c1.metric("Agent Terjadwal", assignments_df["Agent ID"].nunique() if not assignments_df.empty else 0)
            c2.metric("Hari Libur", len(st.session_state.get("holidays", {})))
            c3.metric("Periode Poster", len(weeks_state))

            if not st.session_state.get("validation_ok"):
                for msg in st.session_state.get("validation_messages", []):
                    st.error(msg)

            # Daftar agen bulanan ditampilkan sebelum poster mingguan. User dapat
            # menentukan kolom yang ingin ditampilkan dan urutan Total Kehadiran.
            roster_agents = st.session_state.get("agents", []) or st.session_state.get("current_agents", [])
            generated_target = st.session_state.get("generated_target") or (selected_year, selected_month)
            roster_year, roster_month = generated_target

            roster_control_1, roster_control_2 = st.columns([1.4, 1], gap="large")
            with roster_control_1:
                selected_roster_columns = st.multiselect(
                    "Kolom yang ditampilkan",
                    options=list(ROSTER_COLUMN_OPTIONS),
                    default=list(ROSTER_COLUMN_OPTIONS),
                    key=f"roster_columns_{roster_year}_{roster_month}",
                    help="Contoh: pilih Nama, Unit, dan Total Kehadiran jika Jabatan tidak perlu ditampilkan.",
                )
            with roster_control_2:
                roster_sort_mode = st.selectbox(
                    "Urutan Total Kehadiran",
                    options=[
                        "Default (sesuai file)",
                        "Terbanyak ke terendah",
                        "Terendah ke terbanyak",
                    ],
                    key=f"roster_sort_{roster_year}_{roster_month}",
                )

            if not selected_roster_columns:
                st.warning("Pilih minimal satu kolom untuk menampilkan daftar agen.")
            elif roster_agents:
                roster_image = render_monthly_agent_roster_image(
                    agents=roster_agents,
                    target_year=roster_year,
                    target_month=roster_month,
                    selected_columns=selected_roster_columns,
                    sort_mode=roster_sort_mode,
                    hub_name=hub_name,
                    logo_text=logo_text,
                )
                st.session_state["monthly_roster_image"] = roster_image
                st.image(roster_image, use_container_width=True)

            st.divider()
            period_numbers = list(weeks_state)
            if period_numbers:
                selected_period = st.selectbox(
                    "Pilih rentang tanggal",
                    options=period_numbers,
                    format_func=lambda p: week_range_label(weeks_state[p]),
                )
                image = st.session_state.get("week_images", {}).get(selected_period)
                if image is not None:
                    st.image(image, use_container_width=True)
        render_credit()

    # ---------------------------------------------------------------- DOWNLOAD
    with tabs[3]:
        st.subheader("Download Hasil")
        assignments_df = st.session_state.get("assignments_df", pd.DataFrame())
        if assignments_df.empty:
            st.info("Generate jadwal terlebih dahulu.")
        else:
            target = st.session_state.get("generated_target") or (selected_year, selected_month)
            file_tag = f"{target[0]}-{target[1]:02d}"
            roster_image = st.session_state.get("monthly_roster_image")
            if roster_image is not None:
                dynamic_zip_bytes = build_images_zip(
                    st.session_state.get("week_images", {}),
                    st.session_state.get("weeks", {}),
                    target[0],
                    target[1],
                    monthly_roster_image=roster_image,
                )
                dynamic_pdf_bytes = build_pdf_from_images(
                    st.session_state.get("week_images", {}),
                    monthly_roster_image=roster_image,
                )
                roster_png_buffer = io.BytesIO()
                roster_image.save(roster_png_buffer, format="PNG", dpi=(300, 300), compress_level=2)
                roster_png_bytes = roster_png_buffer.getvalue()
            else:
                dynamic_zip_bytes = st.session_state.get("image_zip_bytes", b"")
                dynamic_pdf_bytes = st.session_state.get("pdf_bytes", b"")
                roster_png_bytes = b""
            c1, c2, c3 = st.columns(3)
            with c1:
                st.download_button("Download Excel", st.session_state.get("excel_bytes", b""), f"floor_time_{file_tag}.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
            with c2:
                st.download_button("Download PNG (ZIP)", dynamic_zip_bytes, f"floor_time_png_{file_tag}.zip", "application/zip", use_container_width=True)
            with c3:
                st.download_button("Download PDF", dynamic_pdf_bytes, f"floor_time_{file_tag}.pdf", "application/pdf", use_container_width=True)
        render_credit()

    # ------------------------------------------------------------------- HELP
    with tabs[4]:
        st.subheader("Panduan")
        st.markdown(
            """
            1. Pilih **bulan jadwal** di sidebar. Data referensi otomatis menggunakan bulan sebelumnya.
            2. Upload Excel/CSV dengan kolom **Nama, Jabatan, Office, Total**.
            3. Atur **Minimum Total Kehadiran** untuk menentukan agen yang memenuhi syarat.
            4. Hilangkan centang **Masuk Jadwal** jika ada agen yang tidak ingin dijadwalkan pada bulan tersebut.
            5. Atur tanggal merah, kapasitas shift, dan batas **1x / 2x / 3x per agent per minggu**, lalu Generate Jadwal Bulanan.
            """
        )
        render_credit()


if __name__ == "__main__":
    main()
